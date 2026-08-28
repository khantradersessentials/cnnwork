"""
Builds a real, multi-sheet Excel report from the JSON + image files that
train.py actually produced. No numbers are invented here — every cell and
every embedded image traces back to a results.json written by a real run.

Usage:
    python -m src.report --runs runs/exp1/results.json runs/exp2/results.json \
        --out report.xlsx
"""
import argparse
import json
import os

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font


def _plot_history(history_list, title, out_path):
    """history_list: list of per-seed history dicts. Real per-epoch data only."""
    fig, axes = plt.subplots(1, 2, figsize=(9, 3.2))
    for h in history_list:
        axes[0].plot(h["epoch"], h["train_loss"], alpha=0.8, label=f"seed {h.get('seed','?')}")
        axes[1].plot(h["epoch"], h["val_acc"], alpha=0.8, label=f"seed {h.get('seed','?')}")
    axes[0].set_title("Train loss"); axes[0].set_xlabel("epoch")
    axes[1].set_title("Validation accuracy"); axes[1].set_xlabel("epoch")
    for ax in axes:
        ax.grid(alpha=0.25)
    if len(history_list) > 1:
        axes[1].legend(fontsize=7)
    fig.suptitle(title, fontsize=10)
    fig.tight_layout()
    fig.savefig(out_path, dpi=130)
    plt.close(fig)


def build_report(run_paths, out_path):
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append([
        "Run", "Arch", "Dataset", "Resolution", "Optimizer", "Pooling", "Epochs",
        "Seeds (n)", "Test Acc Mean", "Test Acc 95% CI", "FLOPs (M)", "Params (M)",
        "Inference Mean (ms)", "Inference 95% CI (ms)", "Defenses", "Eval Attacks",
        "Adversarial Accuracy (mean, per attack)"
    ])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    epoch_sheet = wb.create_sheet("Epoch Data")
    epoch_sheet.append(["Run", "Seed", "Epoch", "Train Loss", "Val Loss", "Val Acc"])
    for cell in epoch_sheet[1]:
        cell.font = Font(bold=True)

    charts_sheet = wb.create_sheet("Charts")
    gradcam_sheet = wb.create_sheet("Grad-CAM")
    failure_sheet = wb.create_sheet("Failure Cases")
    failure_sheet.append(["Run", "Seed", "True Label", "Predicted Label", "Attack", "Confidence", "Image"])
    for cell in failure_sheet[1]:
        cell.font = Font(bold=True)

    chart_row = 1
    gc_row = 1
    failure_row = 2

    for run_path in run_paths:
        with open(run_path) as f:
            data = json.load(f)
        cfg = data["config"]
        run_name = os.path.basename(os.path.dirname(run_path)) or run_path

        adv_acc_str = "; ".join(
            f"{atk}: {stats['mean']*100:.1f}% [CI {stats['ci_low']*100:.1f}-{stats['ci_high']*100:.1f}]"
            for atk, stats in data["adversarial_accuracy"].items()
        ) or "—"

        summary.append([
            run_name, cfg.get("arch"), cfg.get("dataset"), cfg.get("resolution"), cfg.get("optimizer"),
            "spp" if cfg.get("spp") else ("aspp" if cfg.get("aspp") else "none"),
            cfg.get("epochs"), data["test_acc"]["n"],
            f"{data['test_acc']['mean']*100:.2f}%",
            f"{data['test_acc']['ci_low']*100:.2f}% - {data['test_acc']['ci_high']*100:.2f}%",
            f"{data['flops']/1e6:.2f}" if data.get("flops") else "n/a",
            f"{data['params']/1e6:.3f}" if data.get("params") else "n/a",
            f"{data['inference_ms']['mean']:.3f}",
            f"{data['inference_ms']['ci_low']:.3f} - {data['inference_ms']['ci_high']:.3f}",
            ", ".join(cfg.get("defenses", [])) or "none",
            ", ".join(cfg.get("eval_attacks", [])) or "none",
            adv_acc_str,
        ])

        histories = [r["history"] | {"seed": r["seed"]} for r in data["per_seed_results"]]
        for h in histories:
            for i, ep in enumerate(h["epoch"]):
                epoch_sheet.append([run_name, h["seed"], ep, h["train_loss"][i], h["val_loss"][i], h["val_acc"][i]])

        chart_path = f"/tmp/_chart_{run_name}.png"
        _plot_history(histories, f"{run_name} — {cfg.get('arch')} / {cfg.get('dataset')}", chart_path)
        charts_sheet.cell(row=chart_row, column=1, value=run_name).font = Font(bold=True)
        img = XLImage(chart_path)
        img.width, img.height = 560, 200
        charts_sheet.add_image(img, f"A{chart_row+1}")
        chart_row += 14

        for r in data["per_seed_results"]:
            gc_dir = r.get("gradcam_dir")
            if gc_dir and os.path.isdir(gc_dir):
                gradcam_sheet.cell(row=gc_row, column=1,
                                    value=f"{run_name} — seed {r['seed']}").font = Font(bold=True)
                gc_row += 1
                col = 1
                for fname in sorted(os.listdir(gc_dir))[:6]:
                    img = XLImage(os.path.join(gc_dir, fname))
                    img.width, img.height = 100, 100
                    cell_ref = gradcam_sheet.cell(row=gc_row, column=col).coordinate
                    gradcam_sheet.add_image(img, cell_ref)
                    col += 2
                gc_row += 7

            for fc in r.get("failure_cases", []):
                failure_sheet.append([run_name, r["seed"], fc["true_label"], fc["pred_label"],
                                       fc["attack"], f"{fc['confidence']*100:.1f}%", ""])
                if os.path.exists(fc["image_path"]):
                    img = XLImage(fc["image_path"])
                    img.width, img.height = 60, 60
                    failure_sheet.add_image(img, f"G{failure_row}")
                failure_row += 1

    for sheet in (summary, epoch_sheet, failure_sheet):
        for col_cells in sheet.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)

    wb.save(out_path)
    print(f"Report written to {out_path}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--runs", nargs="+", required=True, help="Path(s) to results.json from train.py")
    p.add_argument("--out", default="report.xlsx")
    args = p.parse_args()
    build_report(args.runs, args.out)


if __name__ == "__main__":
    main()
